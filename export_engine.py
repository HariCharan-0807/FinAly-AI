"""
FinAly AI — Data Export Engine
================================
Generates Excel (.xlsx) and PDF reports from filtered transaction data.
All files are generated entirely in memory — nothing touches disk.
"""
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.orm import Session

import models


# ══════════════════════════════════════════════════════════════════════════════
#  Query Helper — shared by both exporters
# ══════════════════════════════════════════════════════════════════════════════
def _query_filtered_transactions(
    db: Session,
    user_id: int,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    category: Optional[str] = None,
    txn_type: Optional[str] = None,
):
    """Return filtered, ordered Transaction rows."""
    q = db.query(models.Transaction).filter(models.Transaction.user_id == user_id)

    if date_from:
        q = q.filter(models.Transaction.date >= date_from)
    if date_to:
        q = q.filter(models.Transaction.date <= date_to)
    if category and category.lower() != "all":
        q = q.filter(models.Transaction.category == category)
    if txn_type and txn_type.lower() != "all":
        q = q.filter(models.Transaction.transaction_type == txn_type)

    return q.order_by(models.Transaction.date.desc()).all()


# ══════════════════════════════════════════════════════════════════════════════
#  Excel (.xlsx) Export
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel(
    db: Session,
    user_id: int,
    user_email: str,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    category: Optional[str] = None,
    txn_type: Optional[str] = None,
) -> bytes:
    """Generate a styled .xlsx workbook and return raw bytes."""
    txns = _query_filtered_transactions(db, user_id, date_from, date_to, category, txn_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    income_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    money_fmt = '₹#,##0.00'

    # ── Title Row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"FinAly AI — Transaction Export for {user_email}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="4F46E5")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Filter Info ───────────────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    filter_parts = []
    if date_from:
        filter_parts.append(f"From: {date_from.strftime('%d-%b-%Y')}")
    if date_to:
        filter_parts.append(f"To: {date_to.strftime('%d-%b-%Y')}")
    if category and category.lower() != "all":
        filter_parts.append(f"Category: {category}")
    if txn_type and txn_type.lower() != "all":
        filter_parts.append(f"Type: {txn_type}")
    ws["A2"].value = "Filters: " + (" | ".join(filter_parts) if filter_parts else "None (all records)")
    ws["A2"].font = Font(name="Calibri", italic=True, color="666666", size=10)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:F3")
    ws["A3"].value = f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')} | Total: {len(txns)} records"
    ws["A3"].font = Font(name="Calibri", italic=True, color="999999", size=9)

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["Date", "Description", "Category", "Type", "Amount (₹)", "Source"]
    col_widths = [16, 40, 16, 12, 18, 14]
    header_row = 5

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Data Rows ─────────────────────────────────────────────────────────────
    total_income = 0.0
    total_expense = 0.0

    for i, t in enumerate(txns, start=header_row + 1):
        date_str = t.date.strftime("%d-%b-%Y") if t.date else ""
        row_data = [
            date_str,
            t.description or "-",
            t.category,
            t.transaction_type,
            round(t.amount, 2),
            t.import_source or "manual",
        ]
        row_fill = income_fill if t.transaction_type == "Income" else expense_fill
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
            if col == 5:
                cell.number_format = money_fmt

        if t.transaction_type == "Income":
            total_income += t.amount
        else:
            total_expense += t.amount

    # ── Summary Row ───────────────────────────────────────────────────────────
    summary_row = header_row + len(txns) + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=11)

    ws.cell(row=summary_row + 1, column=1, value="Total Income:")
    ws.cell(row=summary_row + 1, column=2, value=round(total_income, 2))
    ws.cell(row=summary_row + 1, column=2).number_format = money_fmt
    ws.cell(row=summary_row + 1, column=2).font = Font(color="10B981", bold=True)

    ws.cell(row=summary_row + 2, column=1, value="Total Expenses:")
    ws.cell(row=summary_row + 2, column=2, value=round(total_expense, 2))
    ws.cell(row=summary_row + 2, column=2).number_format = money_fmt
    ws.cell(row=summary_row + 2, column=2).font = Font(color="EF4444", bold=True)

    ws.cell(row=summary_row + 3, column=1, value="Net Balance:")
    ws.cell(row=summary_row + 3, column=2, value=round(total_income - total_expense, 2))
    ws.cell(row=summary_row + 3, column=2).number_format = money_fmt
    ws.cell(row=summary_row + 3, column=2).font = Font(bold=True, size=12)

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  PDF Export
# ══════════════════════════════════════════════════════════════════════════════
def generate_pdf(
    db: Session,
    user_id: int,
    user_email: str,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    category: Optional[str] = None,
    txn_type: Optional[str] = None,
) -> bytes:
    """Generate a styled PDF report and return raw bytes."""
    txns = _query_filtered_transactions(db, user_id, date_from, date_to, category, txn_type)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # ── Title ─────────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#4F46E5"),
        spaceAfter=6,
    )
    elements.append(Paragraph("FinAly AI — Transaction Report", title_style))

    # ── Sub-header ────────────────────────────────────────────────────────────
    sub_style = ParagraphStyle(
        "ExportSub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#666666"),
        spaceAfter=4,
    )
    elements.append(Paragraph(f"User: {user_email}", sub_style))

    filter_parts = []
    if date_from:
        filter_parts.append(f"From: {date_from.strftime('%d-%b-%Y')}")
    if date_to:
        filter_parts.append(f"To: {date_to.strftime('%d-%b-%Y')}")
    if category and category.lower() != "all":
        filter_parts.append(f"Category: {category}")
    if txn_type and txn_type.lower() != "all":
        filter_parts.append(f"Type: {txn_type}")
    elements.append(Paragraph(
        f"Filters: {' | '.join(filter_parts) if filter_parts else 'None (all records)'}",
        sub_style,
    ))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')} | Records: {len(txns)}",
        sub_style,
    ))
    elements.append(Spacer(1, 10))

    # ── Transaction Table ─────────────────────────────────────────────────────
    table_data = [["Date", "Description", "Category", "Type", "Amount (₹)"]]

    total_income = 0.0
    total_expense = 0.0

    for t in txns:
        date_str = t.date.strftime("%d-%b-%Y") if t.date else ""
        desc = (t.description or "-")[:35]
        amt = f"₹{t.amount:,.2f}"
        table_data.append([date_str, desc, t.category, t.transaction_type, amt])

        if t.transaction_type == "Income":
            total_income += t.amount
        else:
            total_expense += t.amount

    # Summary rows
    table_data.append(["", "", "", "", ""])
    table_data.append(["", "", "", "Total Income:", f"₹{total_income:,.2f}"])
    table_data.append(["", "", "", "Total Expenses:", f"₹{total_expense:,.2f}"])
    table_data.append(["", "", "", "Net Balance:", f"₹{total_income - total_expense:,.2f}"])

    col_widths = [68, 170, 70, 60, 80]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # Date
        ("ALIGN", (3, 1), (3, -1), "CENTER"),  # Type
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),   # Amount
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),

        # Gridlines
        ("GRID", (0, 0), (-1, -5), 0.5, colors.HexColor("#E2E8F0")),

        # Alternate row shading
        ("ROWBACKGROUNDS", (0, 1), (-1, -5), [colors.white, colors.HexColor("#F8FAFC")]),

        # Summary section
        ("FONTNAME", (3, -3), (4, -1), "Helvetica-Bold"),
        ("FONTSIZE", (3, -1), (4, -1), 10),
        ("LINEABOVE", (3, -3), (4, -3), 1, colors.HexColor("#4F46E5")),
    ])

    # Colour-code Income/Expense rows
    for i, t in enumerate(txns, start=1):
        if t.transaction_type == "Income":
            table_style.add("TEXTCOLOR", (4, i), (4, i), colors.HexColor("#10B981"))
        else:
            table_style.add("TEXTCOLOR", (4, i), (4, i), colors.HexColor("#EF4444"))

    table.setStyle(table_style)
    elements.append(table)

    # ── Footer ────────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        "ExportFooter",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#999999"),
        alignment=1,  # center
    )
    elements.append(Paragraph(
        "This report was generated by FinAly AI — Your Intelligent Financial Assistant. "
        "Data is encrypted at rest with AES-256.",
        footer_style,
    ))

    doc.build(elements)
    return buf.getvalue()
